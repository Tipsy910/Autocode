from django.contrib import admin

# Register your models here.
# users/admin.py
import csv, io, secrets
from django.contrib import admin, messages
from django.shortcuts import redirect
from django.urls import path
from django.db import transaction
from openpyxl import load_workbook

from .models import User, Students, Teachers
from .forms import UserImportForm


# ---------- Helpers ----------
def as_clean_str(v):
    """แปลงค่าจาก Excel/CSV ให้เป็นสตริงที่สะอาด
    - 4.0 -> "4"
    - None -> ""
    - อื่น ๆ -> str(v)
    """
    if v is None:
        return ""
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else f"{v:.15g}"
    return str(v)


@admin.register(User)
class UserAdmin(admin.ModelAdmin):
    list_display  = ("email", "first_name", "last_name","personal_id", "role", "is_active", "is_staff")
    list_filter   = ("role", "is_active", "is_staff")
    search_fields = ("email", "personal_id", "first_name", "last_name")
    ordering      = ("email",)
    change_list_template = "admin/users/user/change_list.html"  # ฟอร์ม Import ฝังบนหน้า Users

    # --------- เพิ่ม URL สำหรับ import ---------
    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path("import/", self.admin_site.admin_view(self.import_view), name="users_user_import"),
        ]
        return my_urls + urls

    # --------- อ่านไฟล์เป็น rows:list[dict] ---------
    def _read_rows(self, f):
        if f.name.lower().endswith(".xlsx"):
            wb = load_workbook(filename=f, data_only=True)
            ws = wb.active
            headers = [as_clean_str(c.value).strip().lower() for c in ws[1]]
            rows = []
            for r in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in r):
                    continue
                rows.append({headers[i]: as_clean_str(v) for i, v in enumerate(r)})
            return rows
        elif f.name.lower().endswith(".csv"):
            data = f.read().decode("utf-8")
            reader = csv.DictReader(io.StringIO(data))
            return [{k.strip().lower(): as_clean_str(v) for k, v in row.items()} for row in reader]
        else:
            raise ValueError("รองรับเฉพาะ .xlsx หรือ .csv")

    def _split_name(self, name: str):
        name = (name or "").strip()
        if not name:
            return "", ""
        return (name, "") if " " not in name else name.split(" ", 1)

    def _norm_email(self, email: str):
        return (email or "").replace(" ", "").strip().lower()

    def _map_role(self, role: str):
        r = (role or "").strip().upper()
        return r if r in {"STUDENT", "TEACHER", "ADMIN"} else None

    # --------- ตัวหลัก: Import ----------
    @transaction.atomic
    def import_view(self, request):
        if request.method != "POST":
            return redirect("admin:users_user_changelist")

        form = UserImportForm(request.POST, request.FILES)
        if not form.is_valid():
            self.message_user(request, "ฟอร์มไม่ถูกต้อง", level=messages.ERROR)
            return redirect("admin:users_user_changelist")

        dry_run = form.cleaned_data.get("dry_run", False)

        # อ่านไฟล์
        try:
            rows = self._read_rows(form.cleaned_data["file"])
        except Exception as e:
            self.message_user(request, f"อ่านไฟล์ไม่สำเร็จ: {e}", level=messages.ERROR)
            return redirect("admin:users_user_changelist")

        created = updated = skipped = 0
        errors = []

        for idx, row in enumerate(rows, start=2):
            email = self._norm_email(row.get("email"))
            password_raw = (row.get("password") or "").strip()   # จะตั้งตามนี้ (hash ใน set_password)
            full_name = (row.get("name") or "").strip()
            role_raw = self._map_role(row.get("role"))
            # รับ ID ได้หลายชื่อคอลัมน์ → เก็บลง personal_id
            personal_id = (row.get("id") or row.get("ID") or row.get("personal_id") or "").strip()
            personal_id = as_clean_str(personal_id)

            if not email or not email.endswith("@ubu.ac.th"):
                errors.append(f"แถว {idx}: email ไม่ใช่ @ubu.ac.th -> {email}")
                skipped += 1
                continue
            if role_raw is None:
                errors.append(f"แถว {idx}: role ไม่ถูกต้อง -> {row.get('role')}")
                skipped += 1
                continue

            first_name, last_name = self._split_name(full_name)

            try:
                if User.objects.filter(email=email).exists():
                    # ----- อัปเดต -----
                    u = User.objects.get(email=email)
                    u.first_name, u.last_name = first_name, last_name
                    u.role = role_raw                         # จะไม่ถูกทับเป็น ADMIN แล้ว
                    if personal_id:
                        u.personal_id = personal_id
                    if password_raw:
                        u.set_password(password_raw)
                    if not dry_run:
                        u.save()
                    updated += 1
                else:
                    # ----- สร้างใหม่ -----
                    if not password_raw:
                        password_raw = secrets.token_urlsafe(8)  # ถ้าไม่ได้ให้มา จะสุ่ม
                    u = User(email=email, is_active=True, role=role_raw)
                    u.first_name, u.last_name = first_name, last_name
                    if personal_id:
                        u.personal_id = personal_id
                    u.set_password(password_raw)                # hash password
                    if not dry_run:
                        u.save()
                    created += 1

                # ----- โปรไฟล์ตาม role (ไม่มี student_id/teacher_id แล้ว) -----
                if not dry_run:
                    if role_raw == "STUDENT":
                        Teachers.objects.filter(user=u).delete()
                        Students.objects.update_or_create(user=u, defaults={"name": full_name})
                    elif role_raw == "TEACHER":
                        Students.objects.filter(user=u).delete()
                        Teachers.objects.update_or_create(user=u, defaults={"name": full_name})
                    else:  # ADMIN
                        Students.objects.filter(user=u).delete()
                        Teachers.objects.filter(user=u).delete()

            except Exception as e:
                errors.append(f"แถว {idx}: {e}")
                skipped += 1

        # สรุปผล
        level = messages.WARNING if dry_run else messages.SUCCESS
        self.message_user(
            request,
            f"{'(DRY RUN) ' if dry_run else ''}สร้าง {created}, อัปเดต {updated}, ข้าม {skipped} | error={len(errors)}",
            level=level,
        )
        for m in errors[:10]:
            self.message_user(request, m, level=messages.ERROR)

        return redirect("admin:users_user_changelist")